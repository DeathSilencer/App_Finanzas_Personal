"""
futuro_routes.py — Handlers para la suite de Plan Financiero al Futuro
Maneja Dashboard Maestro, Cetes, Control TDC Nu, Fondo de Emergencia y Retiro SAT.
"""

import json
from datetime import datetime, date
import openpyxl

from config import PATH_FUTURO
from helpers.excel_helpers import (
    parse_fecha, safe_float, safe_int, load_wb_readonly, load_wb_write
)


def get_next_spotify_date(current_date: date) -> date:
    """Devuelve la fecha del próximo cargo de Spotify (día 12)."""
    if current_date.day < 12:
        return date(current_date.year, current_date.month, 12)
    else:
        next_m = current_date.month + 1 if current_date.month < 12 else 1
        next_y = current_date.year if current_date.month < 12 else current_date.year + 1
        return date(next_y, next_m, 12)


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/futuro  — Cargar todos los datos de Plan_Financiero_Futuro.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def handle_get_futuro(handler):
    try:
        wb = load_wb_readonly(PATH_FUTURO)
        ws_dash = wb['Dashboard Maestro']
        ws_tdc  = wb['Control TDC Nu']

        # 1. Leer Parámetros Maestro
        ingreso_base = safe_float(ws_dash["B5"].value, 5000.0)
        tasa_nu      = safe_float(ws_dash["B6"].value, 0.13)
        tasa_cetes   = safe_float(ws_dash["B7"].value, 0.0645)
        tasa_afore   = safe_float(ws_dash["B8"].value, 0.085)

        pct_p1       = safe_float(ws_dash["B9"].value, 0.05)
        pct_p2       = safe_float(ws_dash["B10"].value, 0.50)
        pct_p7       = safe_float(ws_dash["B11"].value, 0.30)
        pct_p3       = safe_float(ws_dash["B12"].value, 0.10)
        pct_p6       = safe_float(ws_dash["B13"].value, 0.05)

        tdc_limite   = safe_float(ws_dash["B14"].value, 500.0)
        tdc_tasa     = safe_float(ws_dash["B15"].value, 0.999)
        tdc_corte    = safe_int(ws_dash["B16"].value, 23)
        tdc_pago     = safe_int(ws_dash["B17"].value, 3)

        # 2. Distribución de Ingresos
        distribucion = {
            "p1_involuntario": { "pct": pct_p1, "quincenal": ingreso_base * pct_p1, "mensual": ingreso_base * pct_p1 * 2, "destino": "Cetesdirecto (3 Meses)" },
            "p2_basicos":      { "pct": pct_p2, "quincenal": ingreso_base * pct_p2, "mensual": ingreso_base * pct_p2 * 2, "destino": "Nu Cajita Básicos (13%) -> TDC / Débito" },
            "p7_ocio":         { "pct": pct_p7, "quincenal": ingreso_base * pct_p7, "mensual": ingreso_base * pct_p7 * 2, "destino": "Nu Cajita Ocio (13%) -> Débito Directo" },
            "p3_emergencia":   { "pct": pct_p3, "quincenal": ingreso_base * pct_p3, "mensual": ingreso_base * pct_p3 * 2, "destino": "Nu Cajita Emergencia (Meta 3 Meses)" },
            "p6_retiro":       { "pct": pct_p6, "quincenal": ingreso_base * pct_p6, "mensual": ingreso_base * pct_p6 * 2, "destino": "AFORE XXI Banorte (Art. 151 LISR)" }
        }

        # 3. CETES tabla a 25 años
        aporte_cetes_anual = (ingreso_base * pct_p1 * 2) * 12
        cetes_tabla = []
        for yr in range(1, 26):
            fv_cetes = aporte_cetes_anual * (((1 + tasa_cetes)**yr - 1) / tasa_cetes) * (1 + tasa_cetes)
            ahorro_bolsa = aporte_cetes_anual * yr
            cetes_tabla.append({
                "anio": f"Año {yr}",
                "ahorro_bolsa": round(ahorro_bolsa, 2),
                "interes_acumulado": round(fv_cetes - ahorro_bolsa, 2),
                "saldo_total": round(fv_cetes, 2)
            })

        # 4. TDC Nu — leer compras activas (filas 13 a 112)
        compras_tdc = []
        today = datetime.now().date()
        proximo_spotify = get_next_spotify_date(today)
        proximo_spotify_str = proximo_spotify.strftime("%d/%m/%Y")

        tiene_spotify_este_mes = False
        primera_fila_vacia = None

        for r in range(13, 113):
            monto    = ws_tdc.cell(row=r, column=3).value
            conc_val = str(ws_tdc.cell(row=r, column=4).value or "")
            fecha_val= ws_tdc.cell(row=r, column=2).value

            if "spotify" in conc_val.lower():
                # Comprobar si corresponde a este mes en curso
                tiene_spotify_este_mes = True

            if monto is not None and str(monto).strip() != "":
                try:
                    m_val = float(monto)
                    if m_val > 0:
                        compras_tdc.append({
                            "fila":      r,
                            "id":        ws_tdc.cell(row=r, column=1).value or (r - 12),
                            "fecha":     parse_fecha(fecha_val),
                            "monto":     m_val,
                            "concepto":  conc_val,
                            "categoria": str(ws_tdc.cell(row=r, column=5).value or "Básicos"),
                            "tipo":      str(ws_tdc.cell(row=r, column=6).value or "Gasto Diario"),
                            "apartado":  str(ws_tdc.cell(row=r, column=7).value or "Sí (En Cajita)"),
                            "estado":    str(ws_tdc.cell(row=r, column=8).value or "Pendiente")
                        })
                except (ValueError, TypeError):
                    pass
            else:
                if primera_fila_vacia is None:
                    primera_fila_vacia = r

        # Auto-registro de Spotify ÚNICAMENTE cuando hoy sea el día 12 exacto del mes y aún no esté en la tabla
        if today.day == 12 and not tiene_spotify_este_mes and primera_fila_vacia is not None and primera_fila_vacia <= 112:
            fila_spot = primera_fila_vacia
            spot_date_str = today.strftime("%Y-%m-%d")
            item_spotify = {
                "fila": fila_spot, "id": fila_spot - 12,
                "fecha": spot_date_str, "monto": 74.0,
                "concepto": "Suscripción Spotify", "categoria": "Ocio",
                "tipo": "Suscripción Recurrente", "apartado": "Sí (En Cajita)", "estado": "Pendiente"
            }
            compras_tdc.append(item_spotify)
            try:
                wb_write = load_wb_write(PATH_FUTURO)
                ws_w_tdc = wb_write['Control TDC Nu']
                ws_w_tdc.cell(row=fila_spot, column=1).value = fila_spot - 12
                ws_w_tdc.cell(row=fila_spot, column=2).value = spot_date_str
                ws_w_tdc.cell(row=fila_spot, column=3).value = 74.0
                ws_w_tdc.cell(row=fila_spot, column=4).value = "Suscripción Spotify"
                ws_w_tdc.cell(row=fila_spot, column=5).value = "Ocio"
                ws_w_tdc.cell(row=fila_spot, column=6).value = "Suscripción Recurrente"
                ws_w_tdc.cell(row=fila_spot, column=7).value = "Sí (En Cajita)"
                ws_w_tdc.cell(row=fila_spot, column=8).value = "Pendiente"
                wb_write.save(PATH_FUTURO)
            except PermissionError:
                pass

        tdc_deuda       = sum(c["monto"] for c in compras_tdc if c["estado"] == "Pendiente")
        interes_atraso  = (tdc_deuda * (tdc_tasa / 12)) * 1.16 if tdc_deuda > 0 else 0.0

        # 5. Fondo Emergencia
        aporte_fe_m = (ingreso_base * pct_p3) * 2
        meta_fe     = (ingreso_base * pct_p2 * 2) * 3
        tasa_nu_m   = tasa_nu / 12
        fe_tabla    = []
        saldo_fe    = 0.0
        interes_acum_fe = 0.0

        for m in range(1, 25):
            interes_m = saldo_fe * tasa_nu_m
            saldo_fe  = saldo_fe + aporte_fe_m + interes_m
            interes_acum_fe += interes_m
            fe_tabla.append({
                "mes": f"Mes {m}",
                "ahorro_bolsa": round(aporte_fe_m * m, 2),
                "interes_mes": round(interes_m, 2),
                "interes_acumulado": round(interes_acum_fe, 2),
                "saldo_total": round(saldo_fe, 2),
                "pct_meta": min(100, round((saldo_fe / meta_fe) * 100, 1)) if meta_fe > 0 else 100,
                "estado": "🏆 Meta Cumplida" if saldo_fe >= meta_fe else "En progreso"
            })

        # 6. Retiro & Beneficio Fiscal SAT
        aporte_retiro_anual = (ingreso_base * pct_p6 * 2) * 12
        tasa_isr_sat        = 0.1088
        devolucion_anual_sat= aporte_retiro_anual * tasa_isr_sat
        sat_tabla           = []
        saldo_afore         = 0.0

        for yr in range(1, 26):
            saldo_afore   = (saldo_afore + aporte_retiro_anual) * (1 + tasa_afore)
            ahorro_bolsa  = aporte_retiro_anual * yr
            total_dev_sat = devolucion_anual_sat * yr
            ganancia_neta = (saldo_afore + total_dev_sat) - ahorro_bolsa
            mult          = round((saldo_afore + total_dev_sat) / ahorro_bolsa, 2) if ahorro_bolsa > 0 else 1.0
            sat_tabla.append({
                "anio": f"Año {yr}",
                "ahorro_bolsa": round(ahorro_bolsa, 2),
                "devuelto_sat": round(total_dev_sat, 2),
                "saldo_afore": round(saldo_afore, 2),
                "ganancia_neta": round(ganancia_neta, 2),
                "efecto_mult": mult
            })

        # 7. Ocio y Subcontabilidad Cajita Turbo Nu (Los otros $2,500)
        registros_ocio = []
        pres_ocio = ingreso_base * pct_p7         # $1,500.00
        pres_emergencia = ingreso_base * pct_p3   # $500.00
        pres_retiro = ingreso_base * pct_p6       # $250.00
        pres_cetes = ingreso_base * pct_p1        # $250.00
        pres_total_otros = pres_ocio + pres_emergencia + pres_retiro + pres_cetes # $2,500.00

        cetes_aportado = pres_cetes
        cetes_estado = "Aportado (Cetesdirecto)"
        emergencia_aportado = pres_emergencia
        retiro_aportado = pres_retiro

        if 'Registro Ocio & Cajita Nu' in wb.sheetnames:
            ws_ocio = wb['Registro Ocio & Cajita Nu']
            # Leer aportaciones activas
            if ws_ocio.max_row >= 7:
                cetes_aportado = safe_float(ws_ocio['H4'].value, pres_cetes)
                cetes_estado = str(ws_ocio['H5'].value or "Aportado (Cetesdirecto)")
                emergencia_aportado = safe_float(ws_ocio['H6'].value, pres_emergencia)
                retiro_aportado = safe_float(ws_ocio['H7'].value, pres_retiro)
            
            # Leer bitácora de ocio
            for r in range(11, 111):
                m_val = ws_ocio.cell(row=r, column=4).value
                if m_val is not None and str(m_val).strip() != "":
                    try:
                        m_num = float(m_val)
                        if m_num > 0:
                            registros_ocio.append({
                                "fila": r,
                                "id": ws_ocio.cell(row=r, column=1).value or (r - 10),
                                "fecha": parse_fecha(ws_ocio.cell(row=r, column=2).value),
                                "dia": str(ws_ocio.cell(row=r, column=3).value or ""),
                                "monto": m_num,
                                "categoria": str(ws_ocio.cell(row=r, column=5).value or "🍕 Salidas & Gustos"),
                                "concepto": str(ws_ocio.cell(row=r, column=6).value or ""),
                                "metodo": str(ws_ocio.cell(row=r, column=7).value or "Débito Nu")
                            })
                    except (ValueError, TypeError):
                        pass

        gasto_real_ocio = sum(r["monto"] for r in registros_ocio)
        remanente_ocio = max(0.0, round(pres_ocio - gasto_real_ocio, 2))
        pct_consumido_ocio = round((gasto_real_ocio / pres_ocio) * 100, 1) if pres_ocio > 0 else 0.0

        # Subcontabilidad Cajita Turbo Nu (Fondo de Emergencia + Ocio Disponible + Retiro SAT)
        saldo_en_cajita_ocio = remanente_ocio
        saldo_en_cajita_emergencia = emergencia_aportado
        saldo_en_cajita_retiro = retiro_aportado
        gran_total_cajita_turbo = round(saldo_en_cajita_ocio + saldo_en_cajita_emergencia + saldo_en_cajita_retiro, 2)
        rend_mensual_cajita = round(gran_total_cajita_turbo * (tasa_nu / 12), 2)

        cajita_turbo = {
            "gran_total": gran_total_cajita_turbo,
            "rendimiento_mensual": rend_mensual_cajita,
            "tasa_anual": tasa_nu,
            "porciones": {
                "emergencia": {
                    "monto": saldo_en_cajita_emergencia,
                    "pct": round((saldo_en_cajita_emergencia / gran_total_cajita_turbo * 100), 1) if gran_total_cajita_turbo > 0 else 0,
                    "etiqueta": "Fondo de Emergencia (Intocable)"
                },
                "ocio": {
                    "monto": saldo_en_cajita_ocio,
                    "pct": round((saldo_en_cajita_ocio / gran_total_cajita_turbo * 100), 1) if gran_total_cajita_turbo > 0 else 0,
                    "etiqueta": "Ocio & Salidas (Disponible para gastar)"
                },
                "retiro": {
                    "monto": saldo_en_cajita_retiro,
                    "pct": round((saldo_en_cajita_retiro / gran_total_cajita_turbo * 100), 1) if gran_total_cajita_turbo > 0 else 0,
                    "etiqueta": "Retiro Fiscal SAT (Apartado)"
                }
            }
        }

        otros_fondos = {
            "presupuesto_total": pres_total_otros,
            "ocio": {
                "presupuesto": pres_ocio,
                "gasto_real": gasto_real_ocio,
                "remanente": remanente_ocio,
                "pct_consumido": pct_consumido_ocio
            },
            "emergencia": {
                "presupuesto": pres_emergencia,
                "aportado": emergencia_aportado
            },
            "retiro": {
                "presupuesto": pres_retiro,
                "aportado": retiro_aportado
            },
            "cetes": {
                "presupuesto": pres_cetes,
                "aportado": cetes_aportado,
                "estado": cetes_estado
            },
            "registros_ocio": list(reversed(registros_ocio)),
            "cajita_turbo": cajita_turbo
        }

        handler.send_json({
            "status": "success",
            "config": {
                "ingreso_base": ingreso_base,
                "tasa_nu":      tasa_nu,
                "tasa_cetes":   tasa_cetes,
                "tasa_afore":   tasa_afore,
                "pct_p1":       pct_p1,
                "pct_p2":       pct_p2,
                "pct_p7":       pct_p7,
                "pct_p3":       pct_p3,
                "pct_p6":       pct_p6,
                "tdc_limite":   tdc_limite,
                "tdc_tasa":     tdc_tasa,
                "tdc_corte":    tdc_corte,
                "tdc_pago":     tdc_pago
            },
            "distribucion": distribucion,
            "otros_fondos": otros_fondos,
            "cetes": {
                "aporte_quincenal":  ingreso_base * pct_p1,
                "aporte_anual":      aporte_cetes_anual,
                "tasa_anual":        tasa_cetes,
                "tabla":             cetes_tabla
            },
            "tdc": {
                "limite":               tdc_limite,
                "deuda_actual":         tdc_deuda,
                "disponible":           max(0.0, tdc_limite - tdc_deuda),
                "utilizacion_pct":      round((tdc_deuda / tdc_limite) * 100, 1) if tdc_limite > 0 else 0.0,
                "corte_dia":            tdc_corte,
                "pago_dia":             tdc_pago,
                "interes_atraso":       interes_atraso,
                "proximo_spotify_fecha":proximo_spotify_str,
                "proximo_spotify_monto":74.0,
                "compras":              list(reversed(compras_tdc))
            },
            "fondo_emergencia": {
                "aporte_quincenal": ingreso_base * pct_p3,
                "aporte_mensual":   aporte_fe_m,
                "meta_total":       meta_fe,
                "tasa_anual":       tasa_nu,
                "mes_meta":         14,
                "tabla":            fe_tabla
            },
            "retiro_sat": {
                "aporte_quincenal":  ingreso_base * pct_p6,
                "aporte_mensual":    ingreso_base * pct_p6 * 2,
                "aporte_anual":      aporte_retiro_anual,
                "tasa_isr":          tasa_isr_sat,
                "devolucion_anual":  devolucion_anual_sat,
                "tasa_afore":        tasa_afore,
                "tabla":             sat_tabla
            }
        })
        wb.close()
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Microsoft Excel. Ciérralo para permitir el acceso."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al leer Excel: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/tdc_add  — Agregar compra TDC (filas 13 a 112)
# ─────────────────────────────────────────────────────────────────────────────
def handle_tdc_add(handler, data):
    try:
        wb     = load_wb_write(PATH_FUTURO)
        ws_tdc = wb['Control TDC Nu']

        target_row = None
        for r in range(13, 113):
            if ws_tdc.cell(row=r, column=3).value is None or str(ws_tdc.cell(row=r, column=3).value).strip() == "":
                target_row = r; break
        if target_row is None:
            target_row = 112

        fecha_input = data.get("fecha", datetime.now().strftime("%Y-%m-%d"))
        monto_val   = safe_float(data.get("monto", 0.0))

        ws_tdc.cell(row=target_row, column=1).value = target_row - 12
        ws_tdc.cell(row=target_row, column=2).value = fecha_input
        ws_tdc.cell(row=target_row, column=3).value = monto_val
        ws_tdc.cell(row=target_row, column=4).value = data.get("concepto", "Compra TDC")
        ws_tdc.cell(row=target_row, column=5).value = data.get("categoria", "Básicos")
        ws_tdc.cell(row=target_row, column=6).value = data.get("tipo", "Gasto Diario")
        ws_tdc.cell(row=target_row, column=7).value = data.get("apartado", "Sí (En Cajita)")
        ws_tdc.cell(row=target_row, column=8).value = "Pendiente"

        wb.save(PATH_FUTURO)
        handler.send_json({"status": "success", "fila": target_row, "message": f"Compra de ${monto_val:.2f} guardada en Excel"})
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para guardar."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al guardar compra: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/tdc_edit  — Editar compra TDC
# ─────────────────────────────────────────────────────────────────────────────
def handle_tdc_edit(handler, data):
    try:
        row_num = int(data.get("fila"))
        wb      = load_wb_write(PATH_FUTURO)
        ws_tdc  = wb['Control TDC Nu']

        ws_tdc.cell(row=row_num, column=2).value = data.get("fecha", datetime.now().strftime("%Y-%m-%d"))
        ws_tdc.cell(row=row_num, column=3).value = safe_float(data.get("monto", 0.0))
        ws_tdc.cell(row=row_num, column=4).value = data.get("concepto", "Compra TDC")
        ws_tdc.cell(row=row_num, column=5).value = data.get("categoria", "Básicos")
        ws_tdc.cell(row=row_num, column=6).value = data.get("tipo", "Gasto Diario")
        ws_tdc.cell(row=row_num, column=7).value = data.get("apartado", "Sí (En Cajita)")
        ws_tdc.cell(row=row_num, column=8).value = data.get("estado", "Pendiente")

        wb.save(PATH_FUTURO)
        handler.send_json({"status": "success", "message": f"Compra #{row_num-12} actualizada en Excel"})
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para editar."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al editar compra: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/tdc_delete  — Borrar compra TDC con reindexación (13 a 112)
# ─────────────────────────────────────────────────────────────────────────────
def handle_tdc_delete(handler, data):
    try:
        row_to_delete = int(data.get("fila"))
        wb     = load_wb_write(PATH_FUTURO)
        ws_tdc = wb['Control TDC Nu']

        active_records = []
        for r in range(13, 113):
            m = ws_tdc.cell(row=r, column=3).value
            if m is not None and str(m).strip() != "":
                try:
                    m_val = float(m)
                    if m_val > 0:
                        active_records.append({
                            "original_row": r,
                            "fecha":    ws_tdc.cell(row=r, column=2).value,
                            "monto":    m_val,
                            "concepto": ws_tdc.cell(row=r, column=4).value,
                            "categoria":ws_tdc.cell(row=r, column=5).value,
                            "tipo":     ws_tdc.cell(row=r, column=6).value,
                            "apartado": ws_tdc.cell(row=r, column=7).value,
                            "estado":   ws_tdc.cell(row=r, column=8).value
                        })
                except (ValueError, TypeError):
                    pass

        remaining = [rec for rec in active_records if rec["original_row"] != row_to_delete]

        for r in range(13, 113):
            ws_tdc.cell(row=r, column=1).value = r - 12
            for c in range(2, 9):
                ws_tdc.cell(row=r, column=c).value = None
            ws_tdc.cell(row=r, column=7).value = "Sí (En Cajita)"
            ws_tdc.cell(row=r, column=8).value = "Pendiente"

        for idx, rec in enumerate(remaining):
            curr = 13 + idx
            ws_tdc.cell(row=curr, column=1).value = idx + 1
            ws_tdc.cell(row=curr, column=2).value = rec["fecha"]
            ws_tdc.cell(row=curr, column=3).value = rec["monto"]
            ws_tdc.cell(row=curr, column=4).value = rec["concepto"]
            ws_tdc.cell(row=curr, column=5).value = rec["categoria"]
            ws_tdc.cell(row=curr, column=6).value = rec["tipo"]
            ws_tdc.cell(row=curr, column=7).value = rec["apartado"]
            ws_tdc.cell(row=curr, column=8).value = rec["estado"]

        wb.save(PATH_FUTURO)
        handler.send_json({"status": "success", "message": f"Compra #{row_to_delete-12} eliminada y lista reorganizada."})
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para borrar."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al borrar compra: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/tdc_pay  — Liquidar todas las compras TDC y reiniciar lista limpia
# ─────────────────────────────────────────────────────────────────────────────
def handle_tdc_pay(handler, data):
    try:
        wb     = load_wb_write(PATH_FUTURO)
        ws_tdc = wb['Control TDC Nu']

        # Limpiar todas las filas 13 a 112 para arrancar desde $0.00
        for r in range(13, 113):
            ws_tdc.cell(row=r, column=1).value = r - 12
            for c in range(2, 9):
                ws_tdc.cell(row=r, column=c).value = None
            ws_tdc.cell(row=r, column=7).value = "Sí (En Cajita)"
            ws_tdc.cell(row=r, column=8).value = "Pendiente"

        wb.save(PATH_FUTURO)
        handler.send_json({
            "status": "success",
            "message": "🎉 ¡Tarjeta de crédito liquidada al 100%! La lista se ha reiniciado en $0.00 para tu nuevo ciclo."
        })
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para liquidar y reiniciar la lista."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al liquidar TDC: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/config  — Actualizar parámetros maestros (celdas amarillas B5:B17)
# ─────────────────────────────────────────────────────────────────────────────
def handle_config_futuro(handler, data):
    try:
        wb      = load_wb_write(PATH_FUTURO)
        ws_dash = wb['Dashboard Maestro']

        if "ingreso_base"  in data: ws_dash["B5"].value  = safe_float(data["ingreso_base"])
        if "tasa_nu"       in data: ws_dash["B6"].value  = safe_float(data["tasa_nu"])
        if "tasa_cetes"    in data: ws_dash["B7"].value  = safe_float(data["tasa_cetes"])
        if "tasa_afore"    in data: ws_dash["B8"].value  = safe_float(data["tasa_afore"])
        if "pct_p1"        in data: ws_dash["B9"].value  = safe_float(data["pct_p1"])
        if "pct_p2"        in data: ws_dash["B10"].value = safe_float(data["pct_p2"])
        if "pct_p7"        in data: ws_dash["B11"].value = safe_float(data["pct_p7"])
        if "pct_p3"        in data: ws_dash["B12"].value = safe_float(data["pct_p3"])
        if "pct_p6"        in data: ws_dash["B13"].value = safe_float(data["pct_p6"])
        if "tdc_limite"    in data: ws_dash["B14"].value = safe_float(data["tdc_limite"])
        if "tdc_tasa"      in data: ws_dash["B15"].value = safe_float(data["tdc_tasa"])
        if "tdc_corte"     in data: ws_dash["B16"].value = safe_int(data["tdc_corte"])
        if "tdc_pago"      in data: ws_dash["B17"].value = safe_int(data["tdc_pago"])

        wb.save(PATH_FUTURO)
        handler.send_json({"status": "success", "message": "Parámetros maestros actualizados en Plan_Financiero_Futuro.xlsx"})
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para guardar configuración."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al guardar configuración: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/gasto_ocio  — Registrar gasto en bitácora de ocio
# ─────────────────────────────────────────────────────────────────────────────
def handle_add_gasto_ocio(handler, data):
    try:
        wb = load_wb_write(PATH_FUTURO)
        if 'Registro Ocio & Cajita Nu' not in wb.sheetnames:
            handler.send_json({"status": "error", "message": "Hoja 'Registro Ocio & Cajita Nu' no encontrada en Excel."}, 500)
            return
        ws_ocio = wb['Registro Ocio & Cajita Nu']

        target_row = None
        for r in range(11, 111):
            val = ws_ocio.cell(row=r, column=4).value
            if val is None or str(val).strip() == "" or float(val or 0) == 0:
                target_row = r
                break

        if not target_row:
            handler.send_json({"status": "error", "message": "Límite de 100 registros de ocio alcanzado para la quincena actual."}, 400)
            return

        now = datetime.now()
        fecha_in = data.get("fecha", now.strftime("%Y-%m-%d"))
        dias_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        try:
            d_obj = datetime.strptime(fecha_in, "%Y-%m-%d")
            dia_str = dias_es[d_obj.weekday()]
        except Exception:
            dia_str = dias_es[now.weekday()]

        monto = safe_float(data.get("monto"), 0.0)
        categoria = str(data.get("categoria", "🍕 Salidas & Gustos"))
        concepto = str(data.get("concepto", "Gasto de ocio"))
        metodo = str(data.get("metodo", "Débito Nu"))

        ws_ocio.cell(row=target_row, column=1, value=target_row - 10)
        ws_ocio.cell(row=target_row, column=2, value=fecha_in)
        ws_ocio.cell(row=target_row, column=3, value=dia_str)
        ws_ocio.cell(row=target_row, column=4, value=monto)
        ws_ocio.cell(row=target_row, column=5, value=categoria)
        ws_ocio.cell(row=target_row, column=6, value=concepto)
        ws_ocio.cell(row=target_row, column=7, value=metodo)

        wb.save(PATH_FUTURO)
        wb.close()
        handler.send_json({"status": "success", "message": f"¡Gasto de ocio de ${monto:.2f} registrado en Cajita Nu!"})
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para guardar."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al guardar gasto de ocio: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/delete_gasto_ocio  — Eliminar gasto de ocio con reindexación
# ─────────────────────────────────────────────────────────────────────────────
def handle_delete_gasto_ocio(handler, data):
    try:
        row_del = int(data.get("fila"))
        print(f"[DEBUG DELETE] Solicitado borrar fila: {row_del}")
        wb = load_wb_write(PATH_FUTURO)
        ws_ocio = wb['Registro Ocio & Cajita Nu']

        active = []
        for r in range(11, 111):
            m = ws_ocio.cell(row=r, column=4).value
            if m is not None and str(m).strip() != "":
                try:
                    m_num = float(m)
                    if m_num > 0:
                        active.append({
                            "original_row": r,
                            "fecha": ws_ocio.cell(row=r, column=2).value,
                            "dia": ws_ocio.cell(row=r, column=3).value,
                            "monto": m_num,
                            "categoria": ws_ocio.cell(row=r, column=5).value,
                            "concepto": ws_ocio.cell(row=r, column=6).value,
                            "metodo": ws_ocio.cell(row=r, column=7).value
                        })
                except Exception:
                    pass

        remaining = [rec for rec in active if rec["original_row"] != row_del]

        # Limpiar filas 11 a 110
        for r in range(11, 111):
            ws_ocio.cell(row=r, column=1).value = r - 10
            for c in range(2, 8):
                ws_ocio.cell(row=r, column=c).value = None

        # Reescribir ordenadamente
        for idx, rec in enumerate(remaining):
            curr = 11 + idx
            ws_ocio.cell(row=curr, column=1, value=idx + 1)
            ws_ocio.cell(row=curr, column=2, value=rec["fecha"])
            ws_ocio.cell(row=curr, column=3, value=rec["dia"])
            ws_ocio.cell(row=curr, column=4, value=rec["monto"])
            ws_ocio.cell(row=curr, column=5, value=rec["categoria"])
            ws_ocio.cell(row=curr, column=6, value=rec["concepto"])
            ws_ocio.cell(row=curr, column=7, value=rec["metodo"])

        wb.save(PATH_FUTURO)
        wb.close()
        handler.send_json({"status": "success", "message": f"Gasto #{row_del-10} eliminado correctamente."})
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para editar."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al eliminar gasto: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/aportacion  — Registrar aportación quincenal activa
# ─────────────────────────────────────────────────────────────────────────────
def handle_update_aportacion_futuro(handler, data):
    try:
        wb = load_wb_write(PATH_FUTURO)
        ws_ocio = wb['Registro Ocio & Cajita Nu']
        tipo = data.get("tipo")  # "cetes", "emergencia", "retiro"
        monto = safe_float(data.get("monto"), 0.0)

        if tipo == "cetes":
            ws_ocio['H4'].value = monto
            ws_ocio['H5'].value = "Aportado (Cetesdirecto)" if monto > 0 else "Pendiente"
        elif tipo == "emergencia":
            ws_ocio['H6'].value = monto
        elif tipo == "retiro":
            ws_ocio['H7'].value = monto

        wb.save(PATH_FUTURO)
        wb.close()
        handler.send_json({"status": "success", "message": f"Aportación a {tipo.capitalize()} actualizada: ${monto:.2f}"})
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para editar."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al actualizar aportación: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# POST /api/futuro/cerrar_quincena  — Archivar quincena de los $2,500 en Histórico
# ─────────────────────────────────────────────────────────────────────────────
def handle_cerrar_quincena_futuro(handler, data):
    try:
        wb = load_wb_write(PATH_FUTURO)
        ws_dash = wb['Dashboard Maestro']
        ws_ocio = wb['Registro Ocio & Cajita Nu']
        ws_hist = wb['Histórico Quincenas Futuro']

        ingreso_base = safe_float(ws_dash["B5"].value, 5000.0)
        pct_p7 = safe_float(ws_dash["B11"].value, 0.30)
        pct_p3 = safe_float(ws_dash["B12"].value, 0.10)
        pct_p6 = safe_float(ws_dash["B13"].value, 0.05)
        pct_p1 = safe_float(ws_dash["B9"].value, 0.05)

        pres_ocio = ingreso_base * pct_p7
        pres_emergencia = ingreso_base * pct_p3
        pres_retiro = ingreso_base * pct_p6
        pres_cetes = ingreso_base * pct_p1

        registros = []
        for r in range(11, 111):
            m = ws_ocio.cell(row=r, column=4).value
            if m is not None and str(m).strip() != "":
                try:
                    m_num = float(m)
                    if m_num > 0:
                        registros.append({
                            "id": ws_ocio.cell(row=r, column=1).value or (r - 10),
                            "fecha": parse_fecha(ws_ocio.cell(row=r, column=2).value),
                            "dia": str(ws_ocio.cell(row=r, column=3).value or ""),
                            "monto": m_num,
                            "categoria": str(ws_ocio.cell(row=r, column=5).value or ""),
                            "concepto": str(ws_ocio.cell(row=r, column=6).value or ""),
                            "metodo": str(ws_ocio.cell(row=r, column=7).value or "")
                        })
                except Exception:
                    pass

        gasto_ocio = sum(r["monto"] for r in registros)
        remanente_ocio = max(0.0, round(pres_ocio - gasto_ocio, 2))

        aporte_emergencia = safe_float(ws_ocio['H6'].value, pres_emergencia)
        aporte_retiro = safe_float(ws_ocio['H7'].value, pres_retiro)
        aporte_cetes = safe_float(ws_ocio['H4'].value, pres_cetes)
        total_cajita_cierre = round(remanente_ocio + aporte_emergencia + aporte_retiro, 2)

        now = datetime.now()
        meses_es = {
            "January":"Enero","February":"Febrero","March":"Marzo","April":"Abril",
            "May":"Mayo","June":"Junio","July":"Julio","August":"Agosto",
            "September":"Septiembre","October":"Octubre","November":"Noviembre","December":"Diciembre"
        }
        mes_ingles = now.strftime("%B")
        mes_nombre = meses_es.get(mes_ingles, mes_ingles)

        nombre_periodo = data.get("periodo") or f"{'1ra' if now.day <= 15 else '2da'} Quincena {mes_nombre} {now.year}"
        anio_val = int(data.get("anio", now.year))
        fecha_cierre_val = data.get("fecha_cierre", now.strftime("%Y-%m-%d"))

        target_row = ws_hist.max_row + 1
        for r in range(2, ws_hist.max_row + 2):
            if ws_hist.cell(row=r, column=2).value is None or str(ws_hist.cell(row=r, column=2).value).strip() == "":
                target_row = r
                break

        id_cierre = target_row - 1
        ws_hist.cell(row=target_row, column=1, value=id_cierre)
        ws_hist.cell(row=target_row, column=2, value=nombre_periodo)
        ws_hist.cell(row=target_row, column=3, value=mes_nombre)
        ws_hist.cell(row=target_row, column=4, value=anio_val)
        ws_hist.cell(row=target_row, column=5, value=fecha_cierre_val)
        ws_hist.cell(row=target_row, column=6, value=pres_ocio)
        ws_hist.cell(row=target_row, column=7, value=gasto_ocio)
        ws_hist.cell(row=target_row, column=8, value=remanente_ocio)
        ws_hist.cell(row=target_row, column=9, value=aporte_emergencia)
        ws_hist.cell(row=target_row, column=10, value=aporte_retiro)
        ws_hist.cell(row=target_row, column=11, value=aporte_cetes)
        ws_hist.cell(row=target_row, column=12, value=total_cajita_cierre)
        ws_hist.cell(row=target_row, column=13, value=len(registros))
        detalle_json = json.dumps({
            "registros_ocio": registros,
            "pres_total_otros": 2500.0,
            "gasto_ocio": gasto_ocio,
            "remanente_ocio": remanente_ocio
        }, ensure_ascii=False)
        ws_hist.cell(row=target_row, column=14, value=detalle_json)

        # Resetear bitácora de ocio para la nueva quincena
        for r in range(11, 111):
            ws_ocio.cell(row=r, column=1).value = r - 10
            for c in range(2, 8):
                ws_ocio.cell(row=r, column=c).value = None

        wb.save(PATH_FUTURO)
        wb.close()
        handler.send_json({
            "status": "success",
            "message": f"🎉 ¡Quincena archivada en Plan a Futuro! Remanente de Ocio: ${remanente_ocio:.2f} resguardado en Cajita Turbo."
        })
    except PermissionError:
        handler.send_json({"status": "error", "message": "⚠️ Plan_Financiero_Futuro.xlsx está abierto en Excel. Ciérralo para archivar."}, 423)
    except Exception as e:
        handler.send_json({"status": "error", "message": f"Error al cerrar quincena: {str(e)}"}, 500)


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/futuro/historial  — Obtener quincenas archivadas de Futuro
# ─────────────────────────────────────────────────────────────────────────────
def handle_get_historial_futuro(handler):
    try:
        wb = load_wb_readonly(PATH_FUTURO)
        if 'Histórico Quincenas Futuro' not in wb.sheetnames:
            wb.close()
            handler.send_json({"status": "success", "cierres": []})
            return
        ws_hist = wb['Histórico Quincenas Futuro']
        cierres = []
        for r in range(2, ws_hist.max_row + 1):
            id_val = ws_hist.cell(row=r, column=1).value
            if id_val is not None and str(id_val).strip() != "":
                detalle_raw = ws_hist.cell(row=r, column=14).value or "{}"
                try:
                    detalle_obj = json.loads(str(detalle_raw))
                except Exception:
                    detalle_obj = {}
                cierres.append({
                    "fila": r,
                    "id": int(id_val),
                    "periodo": str(ws_hist.cell(row=r, column=2).value or ""),
                    "mes": str(ws_hist.cell(row=r, column=3).value or ""),
                    "anio": safe_int(ws_hist.cell(row=r, column=4).value, datetime.now().year),
                    "fecha_cierre": str(ws_hist.cell(row=r, column=5).value or ""),
                    "presupuesto_ocio": safe_float(ws_hist.cell(row=r, column=6).value),
                    "gasto_ocio": safe_float(ws_hist.cell(row=r, column=7).value),
                    "remanente_ocio": safe_float(ws_hist.cell(row=r, column=8).value),
                    "aporte_emergencia": safe_float(ws_hist.cell(row=r, column=9).value),
                    "aporte_retiro": safe_float(ws_hist.cell(row=r, column=10).value),
                    "aporte_cetes": safe_float(ws_hist.cell(row=r, column=11).value),
                    "total_cajita_cierre": safe_float(ws_hist.cell(row=r, column=12).value),
                    "num_movimientos": safe_int(ws_hist.cell(row=r, column=13).value),
                    "detalle": detalle_obj
                })
        wb.close()
        handler.send_json({"status": "success", "cierres": list(reversed(cierres))})
    except Exception as e:
        handler.send_json({"status": "error", "message": str(e)}, 500)

